import datetime
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import requests
from paras import *
#look==> https://opendata.cwb.gov.tw/dist/opendata-swagger.html

root_url = 'http://opendata.cwb.gov.tw/api/v1/rest/datastore/'
dataid_list = ['O-A0001-001'] #you can add the others dataid for list

limit = 100
location = '坪林'

def get_data(data_id, limit, location):
	api_url = '{}{}?Authorization={}&locationName={}'.format(root_url, data_id, api_key, location)
	r = requests.get(api_url)
	data = r.json()
	return data

def parse_json(data):
	columns = ['stationId', 'locationName', 'lat', 'lon', 'obstime', 'ELEV', 'WDIR', 'WDSD', 'TEMP', 'HUMD', 'PRES', 'H_24R', 'H_FX', 'H_XD', 'H_FXT', 'D_TX', 'D_TXT', 'D_TN', 'D_TNT', 'CITY', 'CITY_SN', 'TOWN', 'TOWN_SN']
	df = pd.DataFrame(columns=columns)
	data_dict = {}
	locations = data['records']['location']
	row = -1
	#if data include lot of location, then cycle
	for location in locations:
		row = row + 1
		data_dict['stationId'] = location['stationId']
		data_dict['locationName'] = location['locationName']
		data_dict['obstime'] = location['time']['obsTime']
		data_dict['lat'] = location['lat']
		data_dict['lon'] = location['lon']

		factors = location['weatherElement']
		for factor in factors:
			factor_name = factor['elementName']
			data_dict[factor_name] = factor['elementValue']

		parameters = location['parameter']
		for parameter in parameters:
			factor_name = parameter['parameterName']
			data_dict[factor_name] = parameter['parameterValue']

		for key in data_dict.keys():
			df.loc[row,key] = data_dict[key]
	return df

if __name__ == "__main__":
	for dataid in dataid_list:
		json_data = get_data(dataid, limit, location)
		df = parse_json(json_data)

		current_datetime = datetime.datetime.now()
		df['downloadTime'] = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

		#call old .xlsx file and append df on it==>startrow
		book = load_workbook('O-A0001-001坪林data.xlsx')
		writer = pd.ExcelWriter('O-A0001-001坪林data.xlsx', engine='openpyxl')
		writer.book = book
		writer.sheets = {ws.title: ws for ws in book.worksheets}

		for sheetname in writer.sheets:
			df.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)
		writer.save()

		print('save ok:)')
